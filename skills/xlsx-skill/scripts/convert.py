#!/usr/bin/env python3
"""
XLSX 转换脚本
将 Excel 表格转换为 Markdown + 公式 + 截图
"""

import sys
import os
import subprocess
import tempfile
from pathlib import Path


def get_formulas(file_path: Path) -> dict:
    """提取工作表中的公式"""
    from openpyxl import load_workbook

    formulas = {}
    try:
        wb = load_workbook(str(file_path), data_only=False)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_formulas = []
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # 获取计算结果
                        wb_data = load_workbook(str(file_path), data_only=True)
                        result = wb_data[sheet_name][cell.coordinate].value
                        sheet_formulas.append({
                            'cell': cell.coordinate,
                            'formula': cell.value,
                            'result': result
                        })
            if sheet_formulas:
                formulas[sheet_name] = sheet_formulas
        wb.close()
    except Exception as e:
        print(f"警告: 提取公式时出错 - {e}", file=sys.stderr)

    return formulas


def generate_screenshots(file_path: Path, output_dir: Path) -> list:
    """使用 LibreOffice 生成工作表截图"""
    sheets_dir = output_dir / "sheets"
    sheets_dir.mkdir(exist_ok=True)

    screenshots = []

    # 检查 LibreOffice 是否可用
    libreoffice_paths = [
        '/Applications/LibreOffice.app/Contents/MacOS/soffice',
        '/usr/bin/libreoffice',
        '/usr/bin/soffice',
    ]

    soffice = None
    for path in libreoffice_paths:
        if os.path.exists(path):
            soffice = path
            break

    if not soffice:
        print("提示: LibreOffice 未安装，跳过截图生成", file=sys.stderr)
        return screenshots

    try:
        # 转换为 PDF
        with tempfile.TemporaryDirectory() as tmpdir:
            result = subprocess.run([
                soffice,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', tmpdir,
                str(file_path)
            ], capture_output=True, text=True, timeout=60)

            if result.returncode != 0:
                print(f"警告: PDF 转换失败 - {result.stderr}", file=sys.stderr)
                return screenshots

            # 找到生成的 PDF
            pdf_files = list(Path(tmpdir).glob('*.pdf'))
            if not pdf_files:
                return screenshots

            pdf_path = pdf_files[0]

            # PDF 转图片
            try:
                from pdf2image import convert_from_path

                images = convert_from_path(str(pdf_path), dpi=150)
                for i, image in enumerate(images, 1):
                    img_path = sheets_dir / f"sheet_{i:02d}.png"
                    image.save(str(img_path), 'PNG')
                    screenshots.append(img_path)
                    print(f"📷 生成截图: {img_path.name}")
            except Exception as e:
                print(f"警告: 图片转换失败 - {e}", file=sys.stderr)

    except subprocess.TimeoutExpired:
        print("警告: LibreOffice 转换超时", file=sys.stderr)
    except Exception as e:
        print(f"警告: 截图生成失败 - {e}", file=sys.stderr)

    return screenshots


def convert_xlsx(file_path: str) -> None:
    """转换 XLSX 文件为 Markdown + 公式 + 截图"""
    from markitdown import MarkItDown

    input_path = Path(file_path).resolve()

    if not input_path.exists():
        print(f"错误: 文件不存在 - {input_path}", file=sys.stderr)
        sys.exit(1)

    if input_path.suffix.lower() not in ('.xlsx', '.xls'):
        print(f"错误: 不是 Excel 文件 - {input_path}", file=sys.stderr)
        sys.exit(1)

    # 创建输出目录
    output_dir = input_path.parent / f"{input_path.name}.claude"
    output_dir.mkdir(exist_ok=True)

    # 使用 markitdown 转换
    md = MarkItDown()
    result = md.convert(str(input_path))

    # 提取公式
    formulas = get_formulas(input_path)

    # 构建完整内容
    content_parts = [result.text_content]

    if formulas:
        content_parts.append("\n\n---\n\n## 公式汇总\n")
        for sheet_name, sheet_formulas in formulas.items():
            content_parts.append(f"\n### {sheet_name}\n")
            for f in sheet_formulas:
                result_str = f['result'] if f['result'] is not None else '(未计算)'
                content_parts.append(f"- **{f['cell']}**: `{f['formula']}` → {result_str}\n")

    # 保存 Markdown 内容
    content_path = output_dir / "content.md"
    content_path.write_text(''.join(content_parts), encoding='utf-8')

    print(f"✅ 转换完成!")
    print(f"📁 输出目录: {output_dir}")
    print(f"📄 Markdown: {content_path}")

    # 生成截图
    screenshots = generate_screenshots(input_path, output_dir)
    if screenshots:
        print(f"🖼️  截图数量: {len(screenshots)}")


def main():
    if len(sys.argv) < 2:
        print("用法: python convert.py <xlsx文件路径>", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    convert_xlsx(file_path)


if __name__ == "__main__":
    main()
