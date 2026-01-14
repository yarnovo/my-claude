#!/usr/bin/env python3
"""
XLSX è½¬æ¢è„šæœ¬
å°† Excel è¡¨æ ¼è½¬æ¢ä¸º Markdown + å…¬å¼ + æˆªå›¾
"""

import sys
import os
import subprocess
import tempfile
from pathlib import Path


def get_formulas(file_path: Path) -> dict:
    """æå–å·¥ä½œè¡¨ä¸­çš„å…¬å¼"""
    from openpyxl import load_workbook

    formulas = {}
    try:
        wb = load_workbook(str(file_path), data_only=False)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_formulas = []
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # è·å–è®¡ç®—ç»“æœ
                        wb_data = load_workbook(str(file_path), data_only=True)
                        result = wb_data[sheet_name][cell.coordinate].value
                        sheet_formulas.append({
                            'cell': cell.coordinate,
                            'formula': cell.value,
                            'result': result
                        })
            if sheet_formulas:
                formulas[sheet_name] = sheet_formulas
        wb.close()
    except Exception as e:
        print(f"è­¦å‘Š: æå–å…¬å¼æ—¶å‡ºé”™ - {e}", file=sys.stderr)

    return formulas


def generate_screenshots(file_path: Path, output_dir: Path) -> list:
    """ä½¿ç”¨ LibreOffice ç”Ÿæˆå·¥ä½œè¡¨æˆªå›¾"""
    sheets_dir = output_dir / "sheets"
    sheets_dir.mkdir(exist_ok=True)

    screenshots = []

    # æ£€æŸ¥ LibreOffice æ˜¯å¦å¯ç”¨
    libreoffice_paths = [
        '/Applications/LibreOffice.app/Contents/MacOS/soffice',
        '/usr/bin/libreoffice',
        '/usr/bin/soffice',
    ]

    soffice = None
    for path in libreoffice_paths:
        if os.path.exists(path):
            soffice = path
            break

    if not soffice:
        print("æç¤º: LibreOffice æœªå®‰è£…ï¼Œè·³è¿‡æˆªå›¾ç”Ÿæˆ", file=sys.stderr)
        return screenshots

    try:
        # è½¬æ¢ä¸º PDF
        with tempfile.TemporaryDirectory() as tmpdir:
            result = subprocess.run([
                soffice,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', tmpdir,
                str(file_path)
            ], capture_output=True, text=True, timeout=60)

            if result.returncode != 0:
                print(f"è­¦å‘Š: PDF è½¬æ¢å¤±è´¥ - {result.stderr}", file=sys.stderr)
                return screenshots

            # æ‰¾åˆ°ç”Ÿæˆçš„ PDF
            pdf_files = list(Path(tmpdir).glob('*.pdf'))
            if not pdf_files:
                return screenshots

            pdf_path = pdf_files[0]

            # PDF è½¬å›¾ç‰‡
            try:
                from pdf2image import convert_from_path

                images = convert_from_path(str(pdf_path), dpi=150)
                for i, image in enumerate(images, 1):
                    img_path = sheets_dir / f"sheet_{i:02d}.png"
                    image.save(str(img_path), 'PNG')
                    screenshots.append(img_path)
                    print(f"ğŸ“· ç”Ÿæˆæˆªå›¾: {img_path.name}")
            except Exception as e:
                print(f"è­¦å‘Š: å›¾ç‰‡è½¬æ¢å¤±è´¥ - {e}", file=sys.stderr)

    except subprocess.TimeoutExpired:
        print("è­¦å‘Š: LibreOffice è½¬æ¢è¶…æ—¶", file=sys.stderr)
    except Exception as e:
        print(f"è­¦å‘Š: æˆªå›¾ç”Ÿæˆå¤±è´¥ - {e}", file=sys.stderr)

    return screenshots


def convert_xlsx(file_path: str) -> None:
    """è½¬æ¢ XLSX æ–‡ä»¶ä¸º Markdown + å…¬å¼ + æˆªå›¾"""
    from markitdown import MarkItDown

    input_path = Path(file_path).resolve()

    if not input_path.exists():
        print(f"é”™è¯¯: æ–‡ä»¶ä¸å­˜åœ¨ - {input_path}", file=sys.stderr)
        sys.exit(1)

    if input_path.suffix.lower() not in ('.xlsx', '.xls'):
        print(f"é”™è¯¯: ä¸æ˜¯ Excel æ–‡ä»¶ - {input_path}", file=sys.stderr)
        sys.exit(1)

    # åˆ›å»ºè¾“å‡ºç›®å½•
    output_dir = input_path.parent / f"{input_path.name}.claude"
    output_dir.mkdir(exist_ok=True)

    # ä½¿ç”¨ markitdown è½¬æ¢
    md = MarkItDown()
    result = md.convert(str(input_path))

    # æå–å…¬å¼
    formulas = get_formulas(input_path)

    # ç”Ÿæˆæˆªå›¾ï¼ˆå…ˆç”Ÿæˆæˆªå›¾ï¼Œä»¥ä¾¿åœ¨ program-output.md ä¸­å¼•ç”¨ï¼‰
    screenshots = generate_screenshots(input_path, output_dir)

    # æ„å»ºå®Œæ•´å†…å®¹
    content_parts = [f"# Excel: {input_path.name}\n\n"]
    content_parts.append(result.text_content)

    if formulas:
        content_parts.append("\n\n---\n\n## å…¬å¼æ±‡æ€»\n")
        for sheet_name, sheet_formulas in formulas.items():
            content_parts.append(f"\n### {sheet_name}\n")
            for f in sheet_formulas:
                result_str = f['result'] if f['result'] is not None else '(æœªè®¡ç®—)'
                content_parts.append(f"- **{f['cell']}**: `{f['formula']}` â†’ {result_str}\n")

    # æ·»åŠ å·¥ä½œè¡¨æˆªå›¾å¼•ç”¨
    if screenshots:
        content_parts.append("\n\n---\n\n## å·¥ä½œè¡¨é¢„è§ˆ\n\n")
        for i, screenshot in enumerate(screenshots, 1):
            content_parts.append(f"### Sheet {i}\n\n")
            content_parts.append(f"![Sheet {i}](sheets/sheet_{i:02d}.png)\n\n")

    # ä¿å­˜ç¨‹åºè½¬æ¢ç»“æœï¼ˆå¾… AI è¿›ä¸€æ­¥å¤„ç†ï¼‰
    output_path = output_dir / "program-output.md"
    output_path.write_text(''.join(content_parts), encoding='utf-8')

    print(f"âœ… è½¬æ¢å®Œæˆ!")
    print(f"ğŸ“ è¾“å‡ºç›®å½•: {output_dir}")
    print(f"ğŸ“„ ç¨‹åºç»“æœ: {output_path}")
    if screenshots:
        print(f"ğŸ–¼ï¸  æˆªå›¾æ•°é‡: {len(screenshots)}")
    print(f"â³ å¾…å¤„ç†: AI è¯»å–æˆªå›¾ç”Ÿæˆ content.md")


def main():
    if len(sys.argv) < 2:
        print("ç”¨æ³•: python convert.py <xlsxæ–‡ä»¶è·¯å¾„>", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    convert_xlsx(file_path)


if __name__ == "__main__":
    main()
